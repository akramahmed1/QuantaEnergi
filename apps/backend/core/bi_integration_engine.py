"""
BI Integration & Export Engine for QuantaEnergi ETRM/CTRM Platform
Implements Excel/BI integration and export capabilities including:
- Excel export templates
- BI connector integration
- Data export formats
- Report generation
- Dashboard integration
"""

import asyncio
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union, Tuple
from dataclasses import dataclass, field
from enum import Enum
import uuid
import json
import numpy as np
import pandas as pd
from abc import ABC, abstractmethod
import threading
import time
import io
import base64
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import xlsxwriter
import csv
import xml.etree.ElementTree as ET
import yaml
import requests
import aiohttp

logger = logging.getLogger(__name__)

class ExportFormat(Enum):
    """Export format enumeration"""
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"
    XML = "xml"
    PDF = "pdf"
    HTML = "html"
    PARQUET = "parquet"
    AVRO = "avro"

class BIConnectorType(Enum):
    """BI connector types"""
    TABLEAU = "tableau"
    POWER_BI = "power_bi"
    QLIK = "qlik"
    LOOKER = "looker"
    METABASE = "metabase"
    GRAFANA = "grafana"
    KIBANA = "kibana"
    SUPERSET = "superset"

class ReportType(Enum):
    """Report type enumeration"""
    TRADING_SUMMARY = "trading_summary"
    RISK_REPORT = "risk_report"
    PNL_REPORT = "pnl_report"
    POSITION_REPORT = "position_report"
    SETTLEMENT_REPORT = "settlement_report"
    COMPLIANCE_REPORT = "compliance_report"
    MARKET_DATA_REPORT = "market_data_report"
    LOGISTICS_REPORT = "logistics_report"
    CUSTOM = "custom"

@dataclass
class ExportTemplate:
    """Export template definition"""
    template_id: str
    name: str
    description: str = ""
    report_type: ReportType = ReportType.CUSTOM
    export_format: ExportFormat = ExportFormat.EXCEL
    template_data: Dict[str, Any] = field(default_factory=dict)
    fields: List[Dict[str, Any]] = field(default_factory=list)
    filters: List[Dict[str, Any]] = field(default_factory=list)
    aggregations: List[Dict[str, Any]] = field(default_factory=list)
    created_by: str = "system"
    created_at: datetime = field(default_factory=datetime.utcnow)
    metadata: Dict[str, Any] = field(default_factory=dict)

@dataclass
class ExportJob:
    """Export job definition"""
    job_id: str
    template_id: str
    parameters: Dict[str, Any] = field(default_factory=dict)
    filters: Dict[str, Any] = field(default_factory=dict)
    format: ExportFormat = ExportFormat.EXCEL
    status: str = "pending"
    progress: float = 0.0
    file_path: Optional[str] = None
    file_size: Optional[int] = None
    created_at: datetime = field(default_factory=datetime.utcnow)
    completed_at: Optional[datetime] = None
    error_message: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)

@dataclass
class BIConnector:
    """BI connector definition"""
    connector_id: str
    name: str
    connector_type: BIConnectorType
    connection_string: str
    credentials: Dict[str, str] = field(default_factory=dict)
    is_active: bool = True
    last_sync: Optional[datetime] = None
    sync_frequency: int = 3600  # seconds
    created_at: datetime = field(default_factory=datetime.utcnow)
    metadata: Dict[str, Any] = field(default_factory=dict)

@dataclass
class Dashboard:
    """Dashboard definition"""
    dashboard_id: str
    name: str
    description: str = ""
    dashboard_type: str = "custom"
    layout: Dict[str, Any] = field(default_factory=dict)
    widgets: List[Dict[str, Any]] = field(default_factory=list)
    filters: List[Dict[str, Any]] = field(default_factory=list)
    refresh_interval: int = 300  # seconds
    is_public: bool = False
    created_by: str = "system"
    created_at: datetime = field(default_factory=datetime.utcnow)
    metadata: Dict[str, Any] = field(default_factory=dict)

class ExcelExporter:
    """Excel export engine"""
    
    def __init__(self):
        self.templates: Dict[str, ExportTemplate] = {}
        
    def create_template(self, template: ExportTemplate) -> bool:
        """Create export template"""
        try:
            self.templates[template.template_id] = template
            logger.info(f"Created export template: {template.name}")
            return True
        except Exception as e:
            logger.error(f"Error creating export template: {e}")
            return False
    
    def export_to_excel(self, data: List[Dict[str, Any]], template_id: str,
                       output_path: str, parameters: Dict[str, Any] = None) -> bool:
        """Export data to Excel"""
        try:
            if template_id not in self.templates:
                logger.error(f"Template not found: {template_id}")
                return False
            
            template = self.templates[template_id]
            
            # Create workbook
            workbook = xlsxwriter.Workbook(output_path)
            
            # Add worksheet
            worksheet = workbook.add_worksheet(template.name)
            
            # Apply template formatting
            self._apply_template_formatting(worksheet, template)
            
            # Write data
            self._write_data_to_worksheet(worksheet, data, template)
            
            # Add charts if specified
            if template.template_data.get("charts"):
                self._add_charts(workbook, worksheet, data, template)
            
            # Close workbook
            workbook.close()
            
            logger.info(f"Excel export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False
    
    def _apply_template_formatting(self, worksheet, template: ExportTemplate):
        """Apply template formatting to worksheet"""
        try:
            # Header formatting
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'align': 'center',
                'valign': 'vcenter',
                'border': 1
            })
            
            # Data formatting
            data_format = workbook.add_format({
                'align': 'left',
                'valign': 'vcenter',
                'border': 1
            })
            
            # Number formatting
            number_format = workbook.add_format({
                'num_format': '#,##0.00',
                'align': 'right',
                'valign': 'vcenter',
                'border': 1
            })
            
            # Date formatting
            date_format = workbook.add_format({
                'num_format': 'yyyy-mm-dd',
                'align': 'center',
                'valign': 'vcenter',
                'border': 1
            })
            
            # Store formats for later use
            worksheet.formats = {
                'header': header_format,
                'data': data_format,
                'number': number_format,
                'date': date_format
            }
            
        except Exception as e:
            logger.error(f"Error applying template formatting: {e}")
    
    def _write_data_to_worksheet(self, worksheet, data: List[Dict[str, Any]], template: ExportTemplate):
        """Write data to worksheet"""
        try:
            if not data:
                return
            
            # Get field definitions
            fields = template.fields
            if not fields:
                # Use all fields from first data row
                fields = [{"name": key, "type": "string"} for key in data[0].keys()]
            
            # Write headers
            for col, field in enumerate(fields):
                worksheet.write(0, col, field.get("display_name", field["name"]), worksheet.formats["header"])
            
            # Write data rows
            for row, record in enumerate(data, 1):
                for col, field in enumerate(fields):
                    field_name = field["name"]
                    field_type = field.get("type", "string")
                    value = record.get(field_name, "")
                    
                    # Format value based on type
                    if field_type == "number":
                        worksheet.write(row, col, value, worksheet.formats["number"])
                    elif field_type == "date":
                        worksheet.write(row, col, value, worksheet.formats["date"])
                    else:
                        worksheet.write(row, col, value, worksheet.formats["data"])
            
            # Auto-fit columns
            for col, field in enumerate(fields):
                worksheet.set_column(col, col, len(field.get("display_name", field["name"])) + 5)
            
        except Exception as e:
            logger.error(f"Error writing data to worksheet: {e}")
    
    def _add_charts(self, workbook, worksheet, data: List[Dict[str, Any]], template: ExportTemplate):
        """Add charts to worksheet"""
        try:
            charts_config = template.template_data.get("charts", [])
            
            for chart_config in charts_config:
                chart_type = chart_config.get("type", "column")
                chart_title = chart_config.get("title", "Chart")
                data_range = chart_config.get("data_range", "A1:B10")
                position = chart_config.get("position", "E2")
                
                # Create chart
                chart = workbook.add_chart({'type': chart_type})
                chart.set_title({'name': chart_title})
                chart.add_series({
                    'values': f"={worksheet.name}!{data_range}",
                    'categories': f"={worksheet.name}!A1:A{len(data)}"
                })
                
                # Insert chart
                worksheet.insert_chart(position, chart)
            
        except Exception as e:
            logger.error(f"Error adding charts: {e}")

class CSVExporter:
    """CSV export engine"""
    
    def export_to_csv(self, data: List[Dict[str, Any]], output_path: str,
                     delimiter: str = ",", encoding: str = "utf-8") -> bool:
        """Export data to CSV"""
        try:
            if not data:
                return False
            
            # Get field names
            fieldnames = list(data[0].keys())
            
            # Write CSV file
            with open(output_path, 'w', newline='', encoding=encoding) as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=delimiter)
                writer.writeheader()
                writer.writerows(data)
            
            logger.info(f"CSV export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return False

class JSONExporter:
    """JSON export engine"""
    
    def export_to_json(self, data: List[Dict[str, Any]], output_path: str,
                      indent: int = 2, ensure_ascii: bool = False) -> bool:
        """Export data to JSON"""
        try:
            with open(output_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(data, jsonfile, indent=indent, ensure_ascii=ensure_ascii, default=str)
            
            logger.info(f"JSON export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            return False

class XMLExporter:
    """XML export engine"""
    
    def export_to_xml(self, data: List[Dict[str, Any]], output_path: str,
                     root_element: str = "data", item_element: str = "item") -> bool:
        """Export data to XML"""
        try:
            # Create root element
            root = ET.Element(root_element)
            
            # Add items
            for record in data:
                item = ET.SubElement(root, item_element)
                for key, value in record.items():
                    field = ET.SubElement(item, key)
                    field.text = str(value)
            
            # Write XML file
            tree = ET.ElementTree(root)
            tree.write(output_path, encoding='utf-8', xml_declaration=True)
            
            logger.info(f"XML export completed: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to XML: {e}")
            return False

class BIConnectorManager:
    """BI connector management"""
    
    def __init__(self):
        self.connectors: Dict[str, BIConnector] = {}
        self.sync_tasks: Dict[str, asyncio.Task] = {}
        
    def add_connector(self, connector: BIConnector) -> bool:
        """Add BI connector"""
        try:
            self.connectors[connector.connector_id] = connector
            
            # Start sync task if connector is active
            if connector.is_active:
                self._start_sync_task(connector)
            
            logger.info(f"Added BI connector: {connector.name}")
            return True
            
        except Exception as e:
            logger.error(f"Error adding BI connector: {e}")
            return False
    
    def _start_sync_task(self, connector: BIConnector):
        """Start sync task for connector"""
        try:
            task = asyncio.create_task(self._sync_connector(connector))
            self.sync_tasks[connector.connector_id] = task
            
        except Exception as e:
            logger.error(f"Error starting sync task: {e}")
    
    async def _sync_connector(self, connector: BIConnector):
        """Sync connector data"""
        try:
            while connector.is_active:
                # Sync data based on connector type
                if connector.connector_type == BIConnectorType.TABLEAU:
                    await self._sync_tableau(connector)
                elif connector.connector_type == BIConnectorType.POWER_BI:
                    await self._sync_power_bi(connector)
                elif connector.connector_type == BIConnectorType.QLIK:
                    await self._sync_qlik(connector)
                elif connector.connector_type == BIConnectorType.LOOKER:
                    await self._sync_looker(connector)
                elif connector.connector_type == BIConnectorType.METABASE:
                    await self._sync_metabase(connector)
                elif connector.connector_type == BIConnectorType.GRAFANA:
                    await self._sync_grafana(connector)
                elif connector.connector_type == BIConnectorType.KIBANA:
                    await self._sync_kibana(connector)
                elif connector.connector_type == BIConnectorType.SUPERSET:
                    await self._sync_superset(connector)
                
                # Update last sync time
                connector.last_sync = datetime.utcnow()
                
                # Wait for next sync
                await asyncio.sleep(connector.sync_frequency)
                
        except Exception as e:
            logger.error(f"Error syncing connector: {e}")
        finally:
            # Remove sync task
            if connector.connector_id in self.sync_tasks:
                del self.sync_tasks[connector.connector_id]
    
    async def _sync_tableau(self, connector: BIConnector):
        """Sync with Tableau"""
        try:
            # Simulate Tableau sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Tableau: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Tableau: {e}")
    
    async def _sync_power_bi(self, connector: BIConnector):
        """Sync with Power BI"""
        try:
            # Simulate Power BI sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Power BI: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Power BI: {e}")
    
    async def _sync_qlik(self, connector: BIConnector):
        """Sync with Qlik"""
        try:
            # Simulate Qlik sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Qlik: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Qlik: {e}")
    
    async def _sync_looker(self, connector: BIConnector):
        """Sync with Looker"""
        try:
            # Simulate Looker sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Looker: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Looker: {e}")
    
    async def _sync_metabase(self, connector: BIConnector):
        """Sync with Metabase"""
        try:
            # Simulate Metabase sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Metabase: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Metabase: {e}")
    
    async def _sync_grafana(self, connector: BIConnector):
        """Sync with Grafana"""
        try:
            # Simulate Grafana sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Grafana: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Grafana: {e}")
    
    async def _sync_kibana(self, connector: BIConnector):
        """Sync with Kibana"""
        try:
            # Simulate Kibana sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Kibana: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Kibana: {e}")
    
    async def _sync_superset(self, connector: BIConnector):
        """Sync with Superset"""
        try:
            # Simulate Superset sync
            await asyncio.sleep(1)
            logger.info(f"Synced with Superset: {connector.name}")
            
        except Exception as e:
            logger.error(f"Error syncing with Superset: {e}")
    
    def get_connector_status(self) -> Dict[str, Any]:
        """Get connector status"""
        try:
            status = {
                "total_connectors": len(self.connectors),
                "active_connectors": len([c for c in self.connectors.values() if c.is_active]),
                "connectors_by_type": {},
                "sync_tasks": len(self.sync_tasks)
            }
            
            # Group by type
            for connector in self.connectors.values():
                connector_type = connector.connector_type.value
                if connector_type not in status["connectors_by_type"]:
                    status["connectors_by_type"][connector_type] = 0
                status["connectors_by_type"][connector_type] += 1
            
            return status
            
        except Exception as e:
            logger.error(f"Error getting connector status: {e}")
            return {"error": str(e)}

class ReportGenerator:
    """Report generation engine"""
    
    def __init__(self):
        self.templates: Dict[str, ExportTemplate] = {}
        self.reports: Dict[str, Dict[str, Any]] = {}
        
    def create_report_template(self, template: ExportTemplate) -> bool:
        """Create report template"""
        try:
            self.templates[template.template_id] = template
            logger.info(f"Created report template: {template.name}")
            return True
        except Exception as e:
            logger.error(f"Error creating report template: {e}")
            return False
    
    def generate_report(self, template_id: str, data: List[Dict[str, Any]],
                       parameters: Dict[str, Any] = None) -> Dict[str, Any]:
        """Generate report"""
        try:
            if template_id not in self.templates:
                return {"error": "Template not found"}
            
            template = self.templates[template_id]
            
            # Generate report based on template
            report_data = self._process_report_data(data, template, parameters)
            
            # Create report record
            report_id = f"RPT_{uuid.uuid4().hex[:8].upper()}"
            report = {
                "report_id": report_id,
                "template_id": template_id,
                "template_name": template.name,
                "report_type": template.report_type.value,
                "data": report_data,
                "parameters": parameters or {},
                "generated_at": datetime.utcnow().isoformat(),
                "status": "completed"
            }
            
            self.reports[report_id] = report
            
            logger.info(f"Generated report: {report_id}")
            return report
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return {"error": str(e)}
    
    def _process_report_data(self, data: List[Dict[str, Any]], template: ExportTemplate,
                            parameters: Dict[str, Any] = None) -> Dict[str, Any]:
        """Process report data"""
        try:
            processed_data = {
                "summary": {},
                "details": data,
                "charts": [],
                "aggregations": {}
            }
            
            # Apply filters
            if template.filters:
                filtered_data = self._apply_filters(data, template.filters, parameters)
                processed_data["details"] = filtered_data
            
            # Apply aggregations
            if template.aggregations:
                aggregations = self._apply_aggregations(processed_data["details"], template.aggregations)
                processed_data["aggregations"] = aggregations
            
            # Generate summary
            processed_data["summary"] = self._generate_summary(processed_data["details"])
            
            # Generate charts
            if template.template_data.get("charts"):
                processed_data["charts"] = self._generate_charts(processed_data["details"], template.template_data["charts"])
            
            return processed_data
            
        except Exception as e:
            logger.error(f"Error processing report data: {e}")
            return {"error": str(e)}
    
    def _apply_filters(self, data: List[Dict[str, Any]], filters: List[Dict[str, Any]],
                      parameters: Dict[str, Any] = None) -> List[Dict[str, Any]]:
        """Apply filters to data"""
        try:
            filtered_data = data.copy()
            
            for filter_config in filters:
                field = filter_config.get("field", "")
                operator = filter_config.get("operator", "equals")
                value = filter_config.get("value", "")
                
                # Get value from parameters if specified
                if value.startswith("${") and value.endswith("}"):
                    param_name = value[2:-1]
                    if parameters and param_name in parameters:
                        value = parameters[param_name]
                
                # Apply filter
                if operator == "equals":
                    filtered_data = [record for record in filtered_data if record.get(field) == value]
                elif operator == "not_equals":
                    filtered_data = [record for record in filtered_data if record.get(field) != value]
                elif operator == "greater_than":
                    filtered_data = [record for record in filtered_data if record.get(field) > value]
                elif operator == "less_than":
                    filtered_data = [record for record in filtered_data if record.get(field) < value]
                elif operator == "contains":
                    filtered_data = [record for record in filtered_data if str(value) in str(record.get(field, ""))]
                elif operator == "not_contains":
                    filtered_data = [record for record in filtered_data if str(value) not in str(record.get(field, ""))]
                elif operator == "is_empty":
                    filtered_data = [record for record in filtered_data if not record.get(field)]
                elif operator == "is_not_empty":
                    filtered_data = [record for record in filtered_data if record.get(field)]
            
            return filtered_data
            
        except Exception as e:
            logger.error(f"Error applying filters: {e}")
            return data
    
    def _apply_aggregations(self, data: List[Dict[str, Any]], aggregations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Apply aggregations to data"""
        try:
            result = {}
            
            for agg_config in aggregations:
                field = agg_config.get("field", "")
                function = agg_config.get("function", "sum")
                alias = agg_config.get("alias", f"{function}_{field}")
                
                # Extract values
                values = [record.get(field) for record in data if record.get(field) is not None]
                
                if not values:
                    result[alias] = 0
                    continue
                
                # Apply aggregation function
                if function == "sum":
                    result[alias] = sum(values)
                elif function == "avg":
                    result[alias] = sum(values) / len(values)
                elif function == "min":
                    result[alias] = min(values)
                elif function == "max":
                    result[alias] = max(values)
                elif function == "count":
                    result[alias] = len(values)
                elif function == "count_distinct":
                    result[alias] = len(set(values))
                else:
                    result[alias] = 0
            
            return result
            
        except Exception as e:
            logger.error(f"Error applying aggregations: {e}")
            return {}
    
    def _generate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate data summary"""
        try:
            if not data:
                return {}
            
            summary = {
                "total_records": len(data),
                "fields": list(data[0].keys()) if data else [],
                "date_range": {},
                "numeric_fields": {}
            }
            
            # Find date range
            date_fields = []
            for field in summary["fields"]:
                if "date" in field.lower() or "time" in field.lower():
                    date_fields.append(field)
            
            if date_fields:
                dates = []
                for record in data:
                    for field in date_fields:
                        if record.get(field):
                            try:
                                if isinstance(record[field], str):
                                    date_val = datetime.fromisoformat(record[field].replace('Z', '+00:00'))
                                else:
                                    date_val = record[field]
                                dates.append(date_val)
                            except:
                                pass
                
                if dates:
                    summary["date_range"] = {
                        "start": min(dates).isoformat(),
                        "end": max(dates).isoformat()
                    }
            
            # Analyze numeric fields
            for field in summary["fields"]:
                values = [record.get(field) for record in data if isinstance(record.get(field), (int, float))]
                if values:
                    summary["numeric_fields"][field] = {
                        "min": min(values),
                        "max": max(values),
                        "avg": sum(values) / len(values),
                        "sum": sum(values)
                    }
            
            return summary
            
        except Exception as e:
            logger.error(f"Error generating summary: {e}")
            return {}
    
    def _generate_charts(self, data: List[Dict[str, Any]], charts_config: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Generate charts data"""
        try:
            charts = []
            
            for chart_config in charts_config:
                chart_type = chart_config.get("type", "bar")
                title = chart_config.get("title", "Chart")
                x_field = chart_config.get("x_field", "")
                y_field = chart_config.get("y_field", "")
                
                # Extract chart data
                chart_data = {
                    "type": chart_type,
                    "title": title,
                    "x_field": x_field,
                    "y_field": y_field,
                    "data": []
                }
                
                # Group data by x_field
                groups = {}
                for record in data:
                    x_value = record.get(x_field, "")
                    y_value = record.get(y_field, 0)
                    
                    if x_value not in groups:
                        groups[x_value] = 0
                    groups[x_value] += y_value
                
                # Convert to chart format
                for x_value, y_value in groups.items():
                    chart_data["data"].append({
                        "x": x_value,
                        "y": y_value
                    })
                
                charts.append(chart_data)
            
            return charts
            
        except Exception as e:
            logger.error(f"Error generating charts: {e}")
            return []

class DashboardManager:
    """Dashboard management"""
    
    def __init__(self):
        self.dashboards: Dict[str, Dashboard] = {}
        self.widget_types = {
            "chart": self._create_chart_widget,
            "table": self._create_table_widget,
            "metric": self._create_metric_widget,
            "gauge": self._create_gauge_widget,
            "map": self._create_map_widget,
            "text": self._create_text_widget
        }
        
    def create_dashboard(self, dashboard: Dashboard) -> bool:
        """Create dashboard"""
        try:
            self.dashboards[dashboard.dashboard_id] = dashboard
            logger.info(f"Created dashboard: {dashboard.name}")
            return True
        except Exception as e:
            logger.error(f"Error creating dashboard: {e}")
            return False
    
    def _create_chart_widget(self, widget_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create chart widget"""
        try:
            return {
                "type": "chart",
                "title": widget_config.get("title", "Chart"),
                "chart_type": widget_config.get("chart_type", "bar"),
                "data_source": widget_config.get("data_source", ""),
                "x_field": widget_config.get("x_field", ""),
                "y_field": widget_config.get("y_field", ""),
                "position": widget_config.get("position", {"x": 0, "y": 0, "width": 6, "height": 4})
            }
        except Exception as e:
            logger.error(f"Error creating chart widget: {e}")
            return {}
    
    def _create_table_widget(self, widget_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create table widget"""
        try:
            return {
                "type": "table",
                "title": widget_config.get("title", "Table"),
                "data_source": widget_config.get("data_source", ""),
                "columns": widget_config.get("columns", []),
                "position": widget_config.get("position", {"x": 0, "y": 0, "width": 12, "height": 6})
            }
        except Exception as e:
            logger.error(f"Error creating table widget: {e}")
            return {}
    
    def _create_metric_widget(self, widget_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create metric widget"""
        try:
            return {
                "type": "metric",
                "title": widget_config.get("title", "Metric"),
                "value": widget_config.get("value", 0),
                "unit": widget_config.get("unit", ""),
                "trend": widget_config.get("trend", "neutral"),
                "position": widget_config.get("position", {"x": 0, "y": 0, "width": 3, "height": 2})
            }
        except Exception as e:
            logger.error(f"Error creating metric widget: {e}")
            return {}
    
    def _create_gauge_widget(self, widget_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create gauge widget"""
        try:
            return {
                "type": "gauge",
                "title": widget_config.get("title", "Gauge"),
                "value": widget_config.get("value", 0),
                "min": widget_config.get("min", 0),
                "max": widget_config.get("max", 100),
                "unit": widget_config.get("unit", ""),
                "position": widget_config.get("position", {"x": 0, "y": 0, "width": 3, "height": 3})
            }
        except Exception as e:
            logger.error(f"Error creating gauge widget: {e}")
            return {}
    
    def _create_map_widget(self, widget_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create map widget"""
        try:
            return {
                "type": "map",
                "title": widget_config.get("title", "Map"),
                "data_source": widget_config.get("data_source", ""),
                "latitude_field": widget_config.get("latitude_field", ""),
                "longitude_field": widget_config.get("longitude_field", ""),
                "position": widget_config.get("position", {"x": 0, "y": 0, "width": 6, "height": 4})
            }
        except Exception as e:
            logger.error(f"Error creating map widget: {e}")
            return {}
    
    def _create_text_widget(self, widget_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create text widget"""
        try:
            return {
                "type": "text",
                "title": widget_config.get("title", "Text"),
                "content": widget_config.get("content", ""),
                "position": widget_config.get("position", {"x": 0, "y": 0, "width": 6, "height": 2})
            }
        except Exception as e:
            logger.error(f"Error creating text widget: {e}")
            return {}

class BIIntegrationEngine:
    """Main BI integration engine"""
    
    def __init__(self):
        self.excel_exporter = ExcelExporter()
        self.csv_exporter = CSVExporter()
        self.json_exporter = JSONExporter()
        self.xml_exporter = XMLExporter()
        self.bi_connector_manager = BIConnectorManager()
        self.report_generator = ReportGenerator()
        self.dashboard_manager = DashboardManager()
        self.export_jobs: Dict[str, ExportJob] = {}
        
    def create_export_template(self, template: ExportTemplate) -> bool:
        """Create export template"""
        try:
            if template.export_format == ExportFormat.EXCEL:
                return self.excel_exporter.create_template(template)
            else:
                return self.report_generator.create_report_template(template)
                
        except Exception as e:
            logger.error(f"Error creating export template: {e}")
            return False
    
    def export_data(self, template_id: str, data: List[Dict[str, Any]],
                   output_path: str, parameters: Dict[str, Any] = None) -> bool:
        """Export data using template"""
        try:
            # Create export job
            job_id = f"JOB_{uuid.uuid4().hex[:8].upper()}"
            job = ExportJob(
                job_id=job_id,
                template_id=template_id,
                parameters=parameters or {},
                format=ExportFormat.EXCEL  # Default format
            )
            
            self.export_jobs[job_id] = job
            
            # Start export job
            asyncio.create_task(self._execute_export_job(job_id, data, output_path))
            
            return True
            
        except Exception as e:
            logger.error(f"Error exporting data: {e}")
            return False
    
    async def _execute_export_job(self, job_id: str, data: List[Dict[str, Any]], output_path: str):
        """Execute export job"""
        try:
            job = self.export_jobs[job_id]
            job.status = "processing"
            job.progress = 0.0
            
            # Get template
            template = self.excel_exporter.templates.get(job.template_id)
            if not template:
                job.status = "failed"
                job.error_message = "Template not found"
                return
            
            # Export data
            if template.export_format == ExportFormat.EXCEL:
                success = self.excel_exporter.export_to_excel(data, job.template_id, output_path, job.parameters)
            elif template.export_format == ExportFormat.CSV:
                success = self.csv_exporter.export_to_csv(data, output_path)
            elif template.export_format == ExportFormat.JSON:
                success = self.json_exporter.export_to_json(data, output_path)
            elif template.export_format == ExportFormat.XML:
                success = self.xml_exporter.export_to_xml(data, output_path)
            else:
                success = False
            
            if success:
                job.status = "completed"
                job.progress = 100.0
                job.file_path = output_path
                job.completed_at = datetime.utcnow()
            else:
                job.status = "failed"
                job.error_message = "Export failed"
            
        except Exception as e:
            logger.error(f"Error executing export job: {e}")
            job.status = "failed"
            job.error_message = str(e)
    
    def get_export_job_status(self, job_id: str) -> Dict[str, Any]:
        """Get export job status"""
        try:
            if job_id not in self.export_jobs:
                return {"error": "Export job not found"}
            
            job = self.export_jobs[job_id]
            
            return {
                "job_id": job_id,
                "template_id": job.template_id,
                "status": job.status,
                "progress": job.progress,
                "file_path": job.file_path,
                "file_size": job.file_size,
                "created_at": job.created_at.isoformat(),
                "completed_at": job.completed_at.isoformat() if job.completed_at else None,
                "error_message": job.error_message
            }
            
        except Exception as e:
            logger.error(f"Error getting export job status: {e}")
            return {"error": str(e)}
    
    def get_comprehensive_status(self) -> Dict[str, Any]:
        """Get comprehensive BI integration status"""
        try:
            return {
                "export_templates": len(self.excel_exporter.templates),
                "report_templates": len(self.report_generator.templates),
                "export_jobs": len(self.export_jobs),
                "bi_connectors": self.bi_connector_manager.get_connector_status(),
                "dashboards": len(self.dashboard_manager.dashboards),
                "timestamp": datetime.utcnow().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error getting comprehensive status: {e}")
            return {"error": str(e)}

# Global BI integration engine instance
bi_integration_engine = BIIntegrationEngine()

def create_export_template(template: ExportTemplate) -> bool:
    """Create export template"""
    return bi_integration_engine.create_export_template(template)

def export_data(template_id: str, data: List[Dict[str, Any]], output_path: str,
               parameters: Dict[str, Any] = None) -> bool:
    """Export data using template"""
    return bi_integration_engine.export_data(template_id, data, output_path, parameters)

def get_export_job_status(job_id: str) -> Dict[str, Any]:
    """Get export job status"""
    return bi_integration_engine.get_export_job_status(job_id)

def get_bi_integration_status() -> Dict[str, Any]:
    """Get comprehensive BI integration status"""
    return bi_integration_engine.get_comprehensive_status()

def add_bi_connector(connector: BIConnector) -> bool:
    """Add BI connector"""
    return bi_integration_engine.bi_connector_manager.add_connector(connector)

def create_dashboard(dashboard: Dashboard) -> bool:
    """Create dashboard"""
    return bi_integration_engine.dashboard_manager.create_dashboard(dashboard)

def generate_report(template_id: str, data: List[Dict[str, Any]],
                   parameters: Dict[str, Any] = None) -> Dict[str, Any]:
    """Generate report"""
    return bi_integration_engine.report_generator.generate_report(template_id, data, parameters)
